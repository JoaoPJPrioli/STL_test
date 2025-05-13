# cad_collision_analyzer/excel_writer.py

import logging
import numpy as np
from typing import List, Dict, Any, Union
import re

# Attempt import
try:
    import openpyxl
    from openpyxl.utils.exceptions import IllegalCharacterError
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    logging.getLogger("CADAnalyzer.excel_writer").critical(f"Failed to import openpyxl: {e}. Is openpyxl installed?")
    raise ImportError(f"openpyxl import failed in excel_writer: {e}") from e

# Module-specific logger
logger = logging.getLogger("CADAnalyzer.excel_writer")

# Configure logging for Excel writer errors (optional separate file)
excel_error_logger = logging.getLogger("ExcelWriterErrors")
if not excel_error_logger.handlers:
    excel_error_handler = logging.FileHandler('excel_writer_errors.log', mode='a')
    excel_error_formatter = logging.Formatter('%(asctime)s - %(levelname)s - File: %(filename)s - Sheet: %(sheet)s - %(message)s')
    excel_error_handler.setFormatter(excel_error_formatter)
    excel_error_logger.addHandler(excel_error_handler)
    excel_error_logger.setLevel(logging.WARNING)
    excel_error_logger.propagate = False


# Regex for characters disallowed in Excel sheet names
INVALID_SHEET_NAME_CHARS = re.compile(r'[\\*?:/\[\]]')
# Maximum sheet name length
MAX_SHEET_NAME_LENGTH = 31

def _sanitize_sheet_name(name: str) -> str:
    """Removes invalid characters and truncates sheet names for Excel."""
    # Remove invalid characters
    sanitized = INVALID_SHEET_NAME_CHARS.sub('_', name)
    # Truncate to maximum length
    truncated = sanitized[:MAX_SHEET_NAME_LENGTH]
    if truncated != name:
        logger.debug(f"Sanitized sheet name from '{name}' to '{truncated}'")
    return truncated

def _sanitize_cell_value(value: Any) -> Any:
    """Removes illegal XML characters from cell values."""
    if isinstance(value, str):
        # Basic check for common illegal characters (control chars except tab, newline, return)
        # Openpyxl usually handles this, but being explicit can sometimes help.
        # This regex removes characters in the ranges \x00-\x08, \x0B-\x0C, \x0E-\x1F
        cleaned_value = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value)
        if cleaned_value != value:
             logger.debug(f"Sanitized cell value content.")
        return cleaned_value
    return value


class ExcelWriter:
    """
    A class for writing analysis data to an Excel file using openpyxl.

    Handles sheet creation, data writing, and basic sanitization.
    """

    def __init__(self, filename: str):
        """
        Initializes the ExcelWriter with a filename and creates a new workbook.

        Args:
            filename: The path to the Excel file to create/overwrite.
        """
        self.filename = filename
        self.wb: Workbook = openpyxl.Workbook()
        # Remove the default "Sheet" if it exists
        if "Sheet" in self.wb.sheetnames:
            try:
                del self.wb["Sheet"]
            except KeyError: # Should not happen with check, but safety first
                pass
        logger.info(f"Initialized ExcelWriter for file: {self.filename}")

    def _log_excel_error(self, sheet_name: str, message: str, level: int = logging.ERROR, exc_info=False):
        """Helper to log errors with context."""
        log_func = logger.error if level == logging.ERROR else logger.warning
        log_func(f"Excel Error (Sheet: '{sheet_name}'): {message}", exc_info=exc_info)
        # Also log to the dedicated error log file
        excel_error_logger.log(level, message, extra={'filename': self.filename, 'sheet': sheet_name}, exc_info=exc_info)


    def add_metadata_sheet(self, metadata: Dict[str, Any], sheet_name: str = "Metadata"):
        """
        Adds a sheet and writes key-value pairs from the metadata dictionary.

        Args:
            metadata: A dictionary containing metadata parameters and values.
            sheet_name: The name for the metadata sheet. Defaults to "Metadata".
        """
        safe_sheet_name = _sanitize_sheet_name(sheet_name)
        logger.debug(f"Adding metadata sheet: '{safe_sheet_name}'")
        try:
            ws: Worksheet = self.wb.create_sheet(safe_sheet_name)
            # Add header row
            ws.append(["Parameter", "Value"])
            # Add data rows
            for key, value in metadata.items():
                try:
                    # Sanitize key and value before appending
                    safe_key = _sanitize_cell_value(key)
                    safe_value = _sanitize_cell_value(value)
                    # Convert numpy types to standard Python types if necessary
                    if isinstance(safe_value, np.generic):
                        safe_value = safe_value.item()
                    ws.append([safe_key, safe_value])
                except IllegalCharacterError as ice:
                    self._log_excel_error(safe_sheet_name, f"Illegal character error for key '{key}'. Value: '{value}'. Error: {ice}", level=logging.WARNING)
                    ws.append([key, f"ERROR: Value contains illegal characters ({ice})"])
                except Exception as cell_e:
                     self._log_excel_error(safe_sheet_name, f"Unexpected error writing metadata row for key '{key}'. Error: {cell_e}", level=logging.WARNING, exc_info=True)
                     ws.append([key, f"ERROR: Failed to write value ({cell_e})"])

        except Exception as e:
            self._log_excel_error(safe_sheet_name, f"Failed to create or write metadata sheet. Error: {e}", exc_info=True)


    def add_component_names_sheet(self, component_names: List[str], sheet_name: str = "Component List"):
        """
        Adds a sheet listing component names with their corresponding matrix indices.

        Args:
            component_names: A list of component names (order matches matrix rows/cols).
            sheet_name: The name for the component list sheet. Defaults to "Component List".
        """
        safe_sheet_name = _sanitize_sheet_name(sheet_name)
        logger.debug(f"Adding component list sheet: '{safe_sheet_name}'")
        if not component_names:
             logger.warning(f"Component names list is empty. Skipping sheet '{safe_sheet_name}'.")
             return

        try:
            ws: Worksheet = self.wb.create_sheet(safe_sheet_name)
            # Add header row
            ws.append(["Matrix Index", "Component Name"])
            # Add data rows
            for i, name in enumerate(component_names):
                try:
                    safe_name = _sanitize_cell_value(name)
                    ws.append([i, safe_name]) # Index i corresponds to row/col i in matrices
                except IllegalCharacterError as ice:
                    self._log_excel_error(safe_sheet_name, f"Illegal character error for component name at index {i}. Name: '{name}'. Error: {ice}", level=logging.WARNING)
                    ws.append([i, f"ERROR: Name contains illegal characters ({ice})"])
                except Exception as cell_e:
                     self._log_excel_error(safe_sheet_name, f"Unexpected error writing component name row for index {i}. Error: {cell_e}", level=logging.WARNING, exc_info=True)
                     ws.append([i, f"ERROR: Failed to write name ({cell_e})"])

        except Exception as e:
             self._log_excel_error(safe_sheet_name, f"Failed to create or write component list sheet. Error: {e}", exc_info=True)


    def add_matrix_sheet(self, sheet_name: str, matrix: np.ndarray, component_names: List[str]):
        """
        Adds a sheet and writes a matrix with component names as row/column headers.

        Args:
            sheet_name: The desired name for the sheet (will be sanitized).
            matrix: A NumPy array representing the matrix data (e.g., 0, 1, -1).
            component_names: A list of component names corresponding to the matrix dimensions.
        """
        safe_sheet_name = _sanitize_sheet_name(sheet_name)
        logger.debug(f"Adding matrix sheet: '{safe_sheet_name}'")

        if matrix is None or not isinstance(matrix, np.ndarray):
             self._log_excel_error(safe_sheet_name, "Invalid matrix data (None or not a numpy array). Skipping sheet.", level=logging.ERROR)
             return
        if not component_names:
            self._log_excel_error(safe_sheet_name, "Component names list is empty. Cannot write matrix headers. Skipping sheet.", level=logging.ERROR)
            return
        if matrix.shape[0] != len(component_names) or matrix.shape[1] != len(component_names):
             self._log_excel_error(safe_sheet_name, f"Matrix dimensions ({matrix.shape}) do not match number of component names ({len(component_names)}). Skipping sheet.", level=logging.ERROR)
             return

        try:
            ws: Worksheet = self.wb.create_sheet(safe_sheet_name)

            # --- Write Header Row (Column Names) ---
            # First cell is empty, followed by component names
            header_row = [""] + [_sanitize_cell_value(name) for name in component_names]
            try:
                ws.append(header_row)
            except Exception as header_e:
                 # Log error for the header row itself
                 self._log_excel_error(safe_sheet_name, f"Failed to write header row. Error: {header_e}", exc_info=True)
                 # Attempt to continue writing data rows? Or stop? Let's stop for this sheet.
                 return


            # --- Write Data Rows (Row Name + Matrix Data) ---
            for i, row_data in enumerate(matrix):
                # First element is the row header (component name)
                row_header = _sanitize_cell_value(component_names[i])
                # Convert numpy types in row_data to standard Python types
                # and sanitize cell values
                sanitized_row_values = [_sanitize_cell_value(val.item() if isinstance(val, np.generic) else val) for val in row_data]

                # Combine header and data for the row
                full_row = [row_header] + sanitized_row_values
                try:
                    ws.append(full_row)
                except IllegalCharacterError as ice:
                     # Log error for specific row, but try to continue
                     self._log_excel_error(safe_sheet_name, f"Illegal character error writing data row {i} (Component: '{component_names[i]}'). Error: {ice}", level=logging.WARNING)
                     # Optionally write an error message in the row?
                     # ws.append([row_header] + ["ERROR: Illegal characters in data"] * len(row_data))
                except Exception as row_e:
                     # Log error for specific row, but try to continue
                     self._log_excel_error(safe_sheet_name, f"Unexpected error writing data row {i} (Component: '{component_names[i]}'). Error: {row_e}", level=logging.WARNING, exc_info=True)
                     # Optionally write an error message in the row?
                     # ws.append([row_header] + [f"ERROR: {row_e}"] * len(row_data))

        except Exception as e:
             self._log_excel_error(safe_sheet_name, f"Failed to create or write matrix sheet. Error: {e}", exc_info=True)


    def save(self):
        """
        Saves the workbook to the specified filename.

        Logs errors if saving fails.
        """
        logger.info(f"Attempting to save Excel workbook to: {self.filename}")
        try:
            self.wb.save(self.filename)
            logger.info(f"Excel workbook saved successfully.")
        except IOError as e:
            # IOError could be permissions, disk full, etc.
            self._log_excel_error("Workbook Save", f"IOError saving Excel file: {e}", exc_info=True)
            # Re-raise the error so the main script knows saving failed critically
            raise
        except Exception as e:
            # Catch other potential saving errors
            self._log_excel_error("Workbook Save", f"Unexpected error saving Excel file: {e}", exc_info=True)
            # Re-raise the error
            raise

