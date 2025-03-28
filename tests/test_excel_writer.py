import pytest
import numpy as np
import openpyxl # Required for tests too
from pathlib import Path
import logging
import datetime
import os

# --- Module Under Test ---
try:
    from cad_collision_analyzer.excel_writer import ExcelWriter, OPENPYXL_INSTALLED
    Workbook = openpyxl.workbook.Workbook # Get Workbook type for mocking
except ImportError:
    # Allow tests to be skipped gracefully if openpyxl is not installed
    OPENPYXL_INSTALLED = False
    ExcelWriter = None
    Workbook = None


# --- Test Data ---
@pytest.fixture
def sample_metadata() -> dict:
    return {
        "Input File": "test_assembly.step",
        "Timestamp": datetime.datetime(2025, 3, 27, 17, 10, 30), # Use datetime object
        "Number of Components": 3,
        "Linear Deflection": 0.1,
        "Status Message": "Completed Successfully",
        "Complex Value": ["a", 1, {"b": 2}] # Test non-primitive type
    }

@pytest.fixture
def sample_component_names() -> list:
    return ["Part_A", "Component_B", "SubAssembly_C"]

@pytest.fixture
def sample_matrix() -> np.ndarray:
    return np.array([
        [1, 0, 1],
        [0, 0, 0],
        [1, 0, 1]
    ])


# --- Test Cases ---

@pytest.mark.skipif(not OPENPYXL_INSTALLED, reason="Requires openpyxl")
def test_excel_writer_workflow(tmp_path, sample_metadata, sample_component_names, sample_matrix):
    """
    Tests the complete workflow: create, add sheets, save, and verify content.
    """
    filename = tmp_path / "test_output.xlsx"
    matrix_sheet_name = "Test_Matrix_Sheet"
    sanitized_sheet_name = "InvalidSheetName" # Test sanitization

    # 1. Create and write data
    try:
        writer = ExcelWriter(str(filename))
        writer.add_metadata_sheet(sample_metadata)
        writer.add_component_names_sheet(sample_component_names)
        writer.add_matrix_sheet(matrix_sheet_name, sample_matrix, sample_component_names)
        # Add sheet with name needing sanitization
        writer.add_matrix_sheet("InvalidSheetName_[*?]", sample_matrix, sample_component_names)
        writer.save()
    except Exception as e:
        pytest.fail(f"ExcelWriter workflow raised an unexpected exception: {e}")

    # 2. Verify file exists
    assert filename.exists(), "Excel file was not created."
    assert filename.stat().st_size > 0, "Excel file is empty."

    # 3. Load and verify content
    loaded_wb = None # Ensure cleanup happens
    try:
        loaded_wb = openpyxl.load_workbook(filename)

        # Verify sheet names
        expected_sheets = ["Metadata", "Component Names", matrix_sheet_name, sanitized_sheet_name]
        assert set(loaded_wb.sheetnames) == set(expected_sheets)
        assert "Sheet" not in loaded_wb.sheetnames # Default sheet should be removed

        # Verify Metadata sheet
        ws_meta = loaded_wb["Metadata"]
        assert ws_meta.cell(row=1, column=1).value == "Parameter"
        assert ws_meta.cell(row=1, column=2).value == "Value"
        # Check a few key-value pairs (Dict order guaranteed in Python 3.7+)
        assert ws_meta.cell(row=2, column=1).value == "Input File"
        assert ws_meta.cell(row=2, column=2).value == sample_metadata["Input File"]
        assert ws_meta.cell(row=3, column=1).value == "Timestamp"
        assert ws_meta.cell(row=3, column=2).value == sample_metadata["Timestamp"] # Datetime obj
        assert ws_meta.cell(row=4, column=1).value == "Number of Components"
        assert ws_meta.cell(row=4, column=2).value == sample_metadata["Number of Components"] # Int
        assert ws_meta.cell(row=5, column=1).value == "Linear Deflection"
        assert ws_meta.cell(row=5, column=2).value == sample_metadata["Linear Deflection"] # Float
        assert ws_meta.cell(row=7, column=1).value == "Complex Value"
        assert ws_meta.cell(row=7, column=2).value == str(sample_metadata["Complex Value"]) # Check string conversion
        # Check font style (basic check for bold)
        assert ws_meta.cell(row=1, column=1).font.b is True
        assert ws_meta.cell(row=2, column=1).font.b is not True # Data cells shouldn't be bold

        # Verify Component Names sheet
        ws_comp = loaded_wb["Component Names"]
        assert ws_comp.cell(row=1, column=1).value == "Index"
        assert ws_comp.cell(row=1, column=2).value == "Component Name"
        assert ws_comp.cell(row=2, column=1).value == 0
        assert ws_comp.cell(row=2, column=2).value == sample_component_names[0]
        assert ws_comp.cell(row=3, column=1).value == 1
        assert ws_comp.cell(row=3, column=2).value == sample_component_names[1]
        assert ws_comp.cell(row=4, column=1).value == 2
        assert ws_comp.cell(row=4, column=2).value == sample_component_names[2]
        assert ws_comp.cell(row=1, column=1).font.b is True # Header bold check

        # Verify Matrix sheet
        ws_mat = loaded_wb[matrix_sheet_name]
        # Check column headers
        assert ws_mat.cell(row=1, column=1).value is None # A1 should be empty
        assert ws_mat.cell(row=1, column=2).value == sample_component_names[0]
        assert ws_mat.cell(row=1, column=3).value == sample_component_names[1]
        assert ws_mat.cell(row=1, column=4).value == sample_component_names[2]
        assert ws_mat.cell(row=1, column=2).font.b is True # Header bold check
        # Check row headers
        assert ws_mat.cell(row=2, column=1).value == sample_component_names[0]
        assert ws_mat.cell(row=3, column=1).value == sample_component_names[1]
        assert ws_mat.cell(row=4, column=1).value == sample_component_names[2]
        assert ws_mat.cell(row=2, column=1).font.b is True # Header bold check
        # Check matrix data (using sample_matrix values)
        assert ws_mat.cell(row=2, column=2).value == sample_matrix[0, 0] # 1 (int)
        assert ws_mat.cell(row=2, column=3).value == sample_matrix[0, 1] # 0 (int)
        assert ws_mat.cell(row=2, column=4).value == sample_matrix[0, 2] # 1 (int)
        assert ws_mat.cell(row=3, column=2).value == sample_matrix[1, 0] # 0 (int)
        assert ws_mat.cell(row=3, column=3).value == sample_matrix[1, 1] # 0 (int)
        assert ws_mat.cell(row=3, column=4).value == sample_matrix[1, 2] # 0 (int)
        assert ws_mat.cell(row=4, column=2).value == sample_matrix[2, 0] # 1 (int)
        assert ws_mat.cell(row=4, column=3).value == sample_matrix[2, 1] # 0 (int)
        assert ws_mat.cell(row=4, column=4).value == sample_matrix[2, 2] # 1 (int)

        # Verify sanitized sheet name exists and has same content (basic check)
        ws_sanitized = loaded_wb[sanitized_sheet_name]
        assert ws_sanitized.cell(row=2, column=2).value == sample_matrix[0, 0]

    except Exception as e:
        pytest.fail(f"Verification of saved Excel file failed: {e}")
    finally:
        # Cleanup - close workbook if loaded
        if loaded_wb:
            loaded_wb.close()


@pytest.mark.skipif(not OPENPYXL_INSTALLED, reason="Requires openpyxl")
def test_add_matrix_invalid_input(tmp_path, sample_component_names, caplog):
    """Tests that errors are logged for invalid matrix inputs."""
    filename = tmp_path / "invalid_matrix.xlsx"
    writer = ExcelWriter(str(filename))

    # Test non-square matrix
    non_square = np.array([[1, 0], [0, 1], [1, 1]])
    sheet_name_1 = "NonSquare"
    with caplog.at_level(logging.ERROR):
        writer.add_matrix_sheet(sheet_name_1, non_square, sample_component_names)
    assert "Matrix must be a square 2D NumPy array" in caplog.text
    assert sheet_name_1 not in writer.wb.sheetnames

    # Test dimension mismatch
    square_2x2 = np.array([[1, 0], [0, 1]])
    sheet_name_2 = "Mismatch"
    with caplog.at_level(logging.ERROR):
         # Provide 3 names for a 2x2 matrix
        writer.add_matrix_sheet(sheet_name_2, square_2x2, sample_component_names)
    assert "Dimension mismatch" in caplog.text
    assert "Matrix size (2x2)" in caplog.text
    assert "component names (3)" in caplog.text
    assert sheet_name_2 not in writer.wb.sheetnames

    # Test invalid sheet name that becomes empty after sanitize
    sheet_name_3 = "[*?]"
    safe_name_3 = "MatrixSheet" # Default safe name
    with caplog.at_level(logging.WARNING):
         writer.add_matrix_sheet(sheet_name_3, sample_matrix, sample_component_names)
    assert f"Sanitized sheet name from '{sheet_name_3}' to '{safe_name_3}'" in caplog.text
    assert safe_name_3 in writer.wb.sheetnames


@pytest.mark.skipif(not OPENPYXL_INSTALLED or Workbook is None, reason="Requires openpyxl and Workbook type")
def test_save_permission_error(tmp_path, mocker, caplog):
    """Tests logging when saving fails due to permissions."""
    filename = tmp_path / "permission_test.xlsx"

    # Create writer instance
    writer = ExcelWriter(str(filename))
    # Add minimal data so save is attempted
    writer.add_component_names_sheet(["CompA"])

    # Mock the save method of the Workbook class to raise PermissionError
    mock_save = mocker.patch.object(Workbook, 'save', side_effect=PermissionError("Mock permission denied"))
    # mocker.patch('openpyxl.workbook.workbook.Workbook.save', side_effect=PermissionError("Mock permission denied")) # Alternative patch path

    with caplog.at_level(logging.ERROR):
        writer.save()

    # Check that save was called
    mock_save.assert_called_once_with(str(filename))
    # Check that the error was logged
    assert "Permission denied or I/O error" in caplog.text
    assert "Mock permission denied" in caplog.text
    # File might exist briefly before save fails or not at all
    # assert not filename.exists() # This check might be unreliable
